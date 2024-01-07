## Ali Haj
import time
from random import randint
import openpyxl
import pyvisa as visa
from datetime import datetime

import pygal
import numpy as np
# PSU
# power supply ip


rm = visa.ResourceManager('@py')


'''change 'TCPIP0::132.68.57.40::inst0::INSTR' to 'TCPIP0:: [your oscilloscope IP adrees ]::inst0::INSTR' '''
oscilloscope_resource =  'TCPIP0::132.68.57.40::inst0::INSTR'
oscilloscope = rm.open_resource(oscilloscope_resource)  # Increased timeout

# Enable the Digital Voltmeter
oscilloscope.write(":DVM:ENABle ON")
oscilloscope.write(":DVM:MODE DC")


#wave genrator ip
'''change 'TCPIP0::132.68.57.39::inst0::INSTR' to 'TCPIP0::[your WG IP adrees ]::inst0::INSTR''''
WG_address = 'TCPIP0::132.68.57.39::inst0::INSTR'
# Open a connection to the wave genrator
WG = rm.open_resource(WG_address)

# Query instrument information
print("Connected to:", WG.query('*IDN?'))

WG.write("SOURce:FUNCtion PULSE")
square_frequency = 100000       # Replace with your desired frequency in Hz
square_amplitude = 10      # Replace with your desired amplitude in volts
square_duty_cycle = 50

pulse_width = square_duty_cycle/square_frequency/100

WG.write('FUNC:PULSE')
WG.write(f"FREQuency {square_frequency}")
WG.write(f"VOLTage {square_amplitude}")
WG.write(f'PULSE:WIDT {pulse_width}S')


#WG.write(f"FUNCtion:SQUare:DCYCle {square_duty_cycle}")

WG.write('OUTP ON')

vout2=0
workbook = openpyxl.Workbook()
sheet = workbook.active
row=1
starttime=time.time()
while True:
    oscilloscope.write(":AUToscale")
    time.sleep(3)
    vout2 = float(oscilloscope.query(":DVM:CURRent?"))
    sheet.cell(row=row, column=2, value=f"{vout2}")
    T=time.time()-starttime
    sheet.cell(row=row, column=1, value=f"{T}")
    row=row+1
    delta =randint(1, 10)/100
    print(f"duty={square_duty_cycle}")
    print(f"VOUT={vout2}")
    if vout2>48.5 and 97 > square_duty_cycle > 2:
        square_duty_cycle=square_duty_cycle-delta*square_duty_cycle
        pulse_width = square_duty_cycle/square_frequency/100
    if vout2<47.5 and 97 > square_duty_cycle > 2:
        square_duty_cycle=square_duty_cycle+delta*square_duty_cycle
        pulse_width = square_duty_cycle/square_frequency/100
    WG.write(f'PULSE:WIDT {pulse_width}S')

